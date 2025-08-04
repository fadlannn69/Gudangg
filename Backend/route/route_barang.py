from fastapi import APIRouter, Depends, HTTPException, Query, File, Form, UploadFile
from sqlmodel import select, Session
from database import get_session
from model.model_barang import Barang, BarangUpdate, BarangCreate, Histori, HistoriCreate , HistoriRead
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import os
from io import BytesIO
from datetime import datetime, date
from typing import List, Optional
from uuid import uuid4
import shutil
from PIL import Image
from dateutil.relativedelta import relativedelta

barang = APIRouter()


def get_ukuran_order(nama: str) -> int:
    ukuran_order = {
        'XS': 1, 'S': 2, 'M': 3, 'L': 4,
        'XL': 5, '2XL': 6, '3XL': 7, '5XL': 8,
    }
    for ukuran, order in ukuran_order.items():
        if f"({ukuran})" in nama.upper():
            return order
    return 999

@barang.get("/laporan/export-excel")
def export_barang_excel(session: Session = Depends(get_session)):
    barangs = session.exec(select(Barang)).all()
    histori_penjualan = session.exec(select(Histori)).all()  # Tambah ini

    waktu = datetime.now()
    tgl = waktu.date()
    wb = Workbook()

    # Sheet 1: Laporan Barang
    ws1 = wb.active
    ws1.title = f"Laporan Barang"

    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers_barang = [
        "Nama Barang", "Harga Barang(Rp)", "Stok Barang(Pcs)",
        "Lokasi Penempatan Barang", "Waktu Pembelian Barang",
        "Usia Barang", "Jenis Barang", "Barang Terjual (Pcs)"
    ]
    ws1.append(headers_barang)

    for col_num, column_title in enumerate(headers_barang, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(barangs, 2):
        sekarang = datetime.now().date()
        selisih = relativedelta(sekarang, item.waktu)
        usia = f"{selisih.days} Hari, {selisih.months} Bulan, {selisih.years} Tahun"

        row_data = [item.nama, item.harga, item.stok, item.lokasi, item.waktu, usia, item.jenis, item.terjual]
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

    for column_cells in ws1.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws1.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Sheet 2: Histori Penjualan
    ws2 = wb.create_sheet(title="Histori Penjualan")

    headers_histori = ["ID Transaksi Barang" , "Nama Barang" , "Jumlah Barang Terjual" , "Total Harga Barang" , "Waktu Transaksi"]
    ws2.append(headers_histori)

    for col_num, column_title in enumerate(headers_histori, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_num, hist in enumerate(histori_penjualan, 2):
        # Dapatkan nama barang (jika foreign key)
        try:
            barang = session.get(Barang, hist.barang_id)
            nama_barang = barang.nama if barang else "Tidak Ditemukan"
        except:
            nama_barang = "Error"

        row_data = [hist.transaksi_id, nama_barang, hist.terjual ,hist.total_harga, hist.waktujual]
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

    for column_cells in ws2.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Return Excel file
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_barang {tgl}.xlsx"}
    )


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploaded_images")
os.makedirs(UPLOAD_DIR, exist_ok=True)


@barang.post("/tambah")
async def tambah_barang(
    nama: str = Form(...),
    harga: int = Form(...),
    stok: int = Form(...),
    lokasi: str = Form(...),
    waktu: str = Form(...),
    jenis: Optional[str] = Form(None),
    gambar: UploadFile = File(None),
    session: Session = Depends(get_session)
):
    if not nama or harga < 0 or stok < 0 or not lokasi or not waktu:
        raise HTTPException(status_code=400, detail="Data tidak lengkap atau tidak valid")

    if session.exec(select(Barang).where(Barang.nama == nama)).first():
        raise HTTPException(status_code=400, detail="Nama barang sudah ada")

    try:
        waktu_date = datetime.strptime(waktu, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Format tanggal tidak valid, gunakan YYYY-MM-DD")

    filename = None
    if gambar:
        ext = gambar.filename.split(".")[-1].lower()
        if ext not in {"jpg", "jpeg", "png"}:
            raise HTTPException(status_code=400, detail="Format gambar harus jpg/jpeg/png")

        filename = f"{uuid4()}.{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)
        try:
            with open(filepath, "wb") as f:
                gambar.file.seek(0)
                shutil.copyfileobj(gambar.file, f)
            with Image.open(filepath) as img:
                img.thumbnail((800, 800))
                img.save(filepath)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Gagal menyimpan gambar: {str(e)}")

    barang_baru = Barang(
        nama=nama,
        harga=harga,
        stok=stok,
        lokasi=lokasi,
        waktu=waktu_date,
        jenis=jenis,
        gambar=filename
    )
    session.add(barang_baru)
    session.commit()
    session.refresh(barang_baru)
    return barang_baru


@barang.get("/ambil", response_model=List[Barang])
def get_barang(skip: int = Query(0, ge=0), limit: int = Query(10, ge=1), jenis: Optional[str] = None,
               session: Session = Depends(get_session)):
    query = select(Barang)
    if jenis:
        query = query.where(Barang.jenis == jenis)
    result = session.exec(query).all()
    result.sort(key=lambda x: get_ukuran_order(x.nama))
    return result[skip: skip + limit]


@barang.get("/jual", response_model=List[Barang])
def get_barang_terjual(skip: int = Query(0, ge=0), limit: int = Query(10, ge=1), jenis: Optional[str] = None,
                        session: Session = Depends(get_session)):
    query = select(Barang).where(Barang.terjual > 0)
    if jenis:
        query = query.where(Barang.jenis == jenis)
    result = session.exec(query).all()
    result.sort(key=lambda x: get_ukuran_order(x.nama))
    return result[skip: skip + limit]


@barang.put("/{id}/jual")
def jual_barang(
    id: int,
    jumlah: int = Form(...),
    transaksi_id: str = Form(...),  
    session: Session = Depends(get_session)
):
    barang = session.exec(select(Barang).where(Barang.id == id)).first()
    if not barang:
        raise HTTPException(status_code=404, detail="Barang tidak ditemukan")
    if jumlah <= 0:
        raise HTTPException(status_code=400, detail="Jumlah harus lebih dari 0")
    if barang.stok < jumlah:
        raise HTTPException(status_code=400, detail="Stok tidak mencukupi")

    existing = session.exec(
        select(Histori).where(
            Histori.transaksi_id == transaksi_id,
            Histori.barang_id == barang.id
        )
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Barang ini sudah tercatat dalam transaksi")

    barang.stok -= jumlah
    barang.terjual += jumlah
    barang.waktujual = datetime.utcnow().date()
    session.add(barang)

    histori = Histori(
        barang_id=barang.id,
        nama=barang.nama,
        terjual=jumlah,
        harga=barang.harga,
        total_harga=barang.harga * jumlah,
        waktujual=datetime.utcnow(),
        transaksi_id=transaksi_id
    )

    session.add(histori)
    session.commit()
    session.refresh(barang)

    return {
        "detail": f"{jumlah} unit dari '{barang.nama}' berhasil dijual",
        "sisa_stok": barang.stok,
        "total_terjual": barang.terjual,
        "transaksi_id": transaksi_id
    }


@barang.put("/update/{nama}")
def update_barang(nama: str, updated_barang: BarangUpdate, session: Session = Depends(get_session)):
    existing_barang = session.exec(select(Barang).where(Barang.nama == nama)).first()
    if not existing_barang:
        raise HTTPException(status_code=404, detail="Barang tidak ditemukan")

    if updated_barang.harga is not None:
        existing_barang.harga = updated_barang.harga
    if updated_barang.stok is not None:
        existing_barang.stok = updated_barang.stok

    session.add(existing_barang)
    session.commit()
    session.refresh(existing_barang)
    return {"detail": f"Barang '{nama}' berhasil diperbarui"}


@barang.delete("/hapus/{nama}")
def delete_barang(nama: str, session: Session = Depends(get_session)):
    existing_barang = session.exec(select(Barang).where(Barang.nama == nama)).first()
    if not existing_barang:
        raise HTTPException(status_code=404, detail="Barang tak ditemukan")
    session.delete(existing_barang)
    session.commit()
    return {"detail": f"{existing_barang.nama} berhasil dihapus"}

@barang.get("/histori", response_model=List[HistoriRead])
def get_histori(session: Session = Depends(get_session)):
    query = (
        session.query(Histori.id, Histori.terjual, Histori.harga,
                      Histori.total_harga, Histori.waktujual,
                      Barang.nama.label("nama"))  
        .join(Barang, Histori.barang_id == Barang.id)
        .all()
    )

    return [HistoriRead(**row._asdict()) for row in query]

